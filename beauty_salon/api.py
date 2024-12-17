from rest_framework import viewsets, serializers
from rest_framework.decorators import action
from rest_framework.response import Response
from django.db.models import Avg, Count, Max, Min, Sum
from django.contrib.auth.models import User
from .models import Client, Service, Master, Appointment, Review
from .serializers import (
    ClientSerializer, ServiceSerializer, MasterSerializer, 
    AppointmentSerializer, ReviewSerializer
)
from rest_framework import serializers
from django.core.cache import cache
import pyotp
from .permissions import OTPRequired
from rest_framework.permissions import IsAuthenticated
from django.http import HttpResponse
import openpyxl
from openpyxl.styles import Font
from datetime import datetime
from rest_framework import status

class BaseModelViewSet(viewsets.ModelViewSet):
    permission_classes = [IsAuthenticated]
    
    class OTPSerializer(serializers.Serializer):
        code = serializers.CharField()

    @action(detail=False, methods=['POST'], url_path='verify-otp')
    def verify_otp(self, request):
        serializer = self.OTPSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        # Для демонстрации используем фиксированный код
        DEMO_OTP = '123456'
        
        success = False
        if serializer.validated_data['code'] == DEMO_OTP:
            # OTP действителен 30 минут
            cache.set(f'otp_good_{request.user.id}', True, 30 * 60)
            success = True

        return Response({
            'success': success,
            'message': 'OTP verified successfully' if success else 'Invalid OTP'
        })

    @action(detail=False, methods=['GET'], url_path='otp-status')
    def get_otp_status(self, request):
        otp_verified = cache.get(f'otp_good_{request.user.id}', False)
        return Response({
            'otp_verified': otp_verified
        })

    @action(detail=False, methods=['GET'], url_path='check-otp-status')
    def check_otp_status(self, request):
        """Endpoint для отладки OTP статуса"""
        current_user_otp = cache.get(f'otp_good_{request.user.id}')
        
        return Response({
            'user_id': request.user.id,
            'username': request.user.username,
            'current_user_otp': current_user_otp,
            'otp_key': f'otp_good_{request.user.id}'
        })

    def get_queryset(self):
        qs = super().get_queryset()
        # Для моделей Service, Master и Client возвращаем все записи
        if isinstance(self, (ServiceViewSet, MasterViewSet, ClientViewSet)):
            return qs
        # Для Appointment и Review фильтруем по пользователю
        if not self.request.user.is_superuser:  # если не danil
            return qs.filter(client__user=self.request.user)  # только свои записи
        return qs  # для danil возвращаем все

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if self.request.user.is_superuser:
            context['users'] = User.objects.all()
        return context

    def destroy(self, request, *args, **kwargs):
        try:
            print(f"Attempting to delete object with kwargs: {kwargs}")  # Отладка
            print(f"User: {request.user}, is_authenticated: {request.user.is_authenticated}")  # Отладка
            instance = self.get_object()
            print(f"Found instance: {instance}")  # Отладка
            self.perform_destroy(instance)
            print("Instance deleted successfully")  # Отладка
            return Response(status=status.HTTP_204_NO_CONTENT)
        except Exception as e:
            print(f"Error during deletion: {str(e)}")  # Отладка
            return Response(
                {'detail': str(e)}, 
                status=status.HTTP_400_BAD_REQUEST
            )

    def get_permissions(self):
        if self.action in ['update', 'partial_update']:
            return [IsAuthenticated(), OTPRequired()]
        return [IsAuthenticated()]

class ClientViewSet(BaseModelViewSet):
    queryset = Client.objects.all()
    serializer_class = ClientSerializer
    http_method_names = ['get', 'post', 'put', 'delete']
    permission_classes = [IsAuthenticated]

    class StatsSerializer(serializers.Serializer):
        total_clients = serializers.IntegerField()
        clients_with_email = serializers.IntegerField()
        most_popular_service = serializers.CharField(allow_null=True)
        appointments_per_client = serializers.FloatField(allow_null=True)

    @action(detail=False, methods=["GET"], url_path="stats")
    def get_stats(self, request):
        queryset = self.get_queryset()
        if not queryset.exists():
            return Response({
                'total_clients': 0,
                'clients_with_email': 0,
                'most_popular_service': None,
                'appointments_per_client': None
            })

        service_stats = queryset.values('service__name')\
            .annotate(count=Count('service'))\
            .order_by('-count').first()

        stats = {
            'total_clients': queryset.count(),
            'clients_with_email': queryset.exclude(email='').count(),
            'most_popular_service': service_stats['service__name'] if service_stats else None,
            'appointments_per_client': Appointment.objects.filter(
                client__in=queryset
            ).count() / queryset.count() if queryset.exists() else None
        }
        
        serializer = self.StatsSerializer(instance=stats)
        return Response(serializer.data)

    def get_queryset(self):
        return Client.objects.filter(user=self.request.user)

    def perform_create(self, serializer):
        serializer.save(user=self.request.user)

class ServiceViewSet(BaseModelViewSet):
    queryset = Service.objects.all()
    serializer_class = ServiceSerializer
    http_method_names = ['get', 'post', 'put', 'delete']

    def get_permissions(self):
        if self.action == 'destroy':
            return [IsAuthenticated()]
        if self.action == 'create':
            return [IsAuthenticated()]  # Разрешаем создание с базовой авторизацией
        return super().get_permissions()

    class StatsSerializer(serializers.Serializer):
        total_services = serializers.IntegerField()
        avg_price = serializers.FloatField()
        max_price = serializers.DecimalField(max_digits=10, decimal_places=2)
        min_price = serializers.DecimalField(max_digits=10, decimal_places=2)
        avg_duration = serializers.FloatField()
        total_revenue = serializers.DecimalField(max_digits=10, decimal_places=2)

    @action(detail=False, methods=["GET"], url_path="stats")
    def get_stats(self, request):
        queryset = self.get_queryset()
        if not queryset.exists():
            return Response({
                'total_services': 0,
                'avg_price': 0,
                'max_price': 0,
                'min_price': 0,
                'avg_duration': 0,
                'total_revenue': 0
            })
        
        stats = queryset.aggregate(
            total_services=Count("*"),
            avg_price=Avg("price"),
            max_price=Max("price"),
            min_price=Min("price"),
            avg_duration=Avg("duration"),
            total_revenue=Sum("price")
        )
        serializer = self.StatsSerializer(instance=stats)
        return Response(serializer.data)

    def perform_create(self, serializer):
        serializer.save()

    @action(detail=False, methods=['GET'], url_path='export-excel')
    def export_excel(self, request):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Услуги"

        # Заголовки
        headers = ['Название', 'Цена', 'Длительность (мин)']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)

        # Данные
        queryset = self.get_queryset()
        for row, service in enumerate(queryset, 2):
            ws.cell(row=row, column=1, value=service.name)
            ws.cell(row=row, column=2, value=float(service.price))
            ws.cell(row=row, column=3, value=service.duration)

        # Автоатическая ширина колонок
        for column in ws.columns:
            max_length = 0
            column_letter = openpyxl.utils.get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=services.xlsx'
        wb.save(response)
        return response

    def destroy(self, request, *args, **kwargs):
        try:
            instance = self.get_object()
            self.perform_destroy(instance)
            return Response(status=status.HTTP_204_NO_CONTENT)
        except Exception as e:
            print(f"Error deleting service: {str(e)}")  # Для отладки
            return Response(
                {'detail': str(e)}, 
                status=status.HTTP_400_BAD_REQUEST
            )

class MasterViewSet(BaseModelViewSet):
    queryset = Master.objects.all()
    serializer_class = MasterSerializer
    http_method_names = ['get', 'post', 'put', 'delete']

    def get_permissions(self):
        if self.action in ['create', 'update', 'partial_update', 'destroy']:
            return [IsAuthenticated()]
        return super().get_permissions()

    class StatsSerializer(serializers.Serializer):
        total_masters = serializers.IntegerField()
        avg_services_per_master = serializers.FloatField(allow_null=True)
        most_common_specialization = serializers.CharField(allow_null=True)
        busiest_master = serializers.CharField(allow_null=True)

    @action(detail=False, methods=["GET"], url_path="stats")
    def get_stats(self, request):
        queryset = self.get_queryset()
        if not queryset.exists():
            return Response({
                'total_masters': 0,
                'avg_services_per_master': None,
                'most_common_specialization': None,
                'busiest_master': None
            })

        busiest = queryset.annotate(
            appointments_count=Count('appointment')
        ).order_by('-appointments_count').first()

        spec_stats = queryset.values('specialization')\
            .annotate(count=Count('specialization'))\
            .order_by('-count').first()

        stats = {
            'total_masters': queryset.count(),
            'avg_services_per_master': queryset.annotate(
                service_count=Count('services')
            ).aggregate(avg=Avg('service_count'))['avg'],
            'most_common_specialization': spec_stats['specialization'] if spec_stats else None,
            'busiest_master': busiest.name if busiest else None
        }
        
        serializer = self.StatsSerializer(instance=stats)
        return Response(serializer.data)

    @action(detail=False, methods=['GET'], url_path='export-excel')
    def export_excel(self, request):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Мастера"

        headers = ['ФИО', 'Специализация', 'Услуги']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)

        queryset = self.get_queryset()
        for row, master in enumerate(queryset, 2):
            ws.cell(row=row, column=1, value=master.name)
            ws.cell(row=row, column=2, value=master.specialization)
            ws.cell(row=row, column=3, value=', '.join([s.name for s in master.services.all()]))

        for column in ws.columns:
            max_length = 0
            column_letter = openpyxl.utils.get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=masters.xlsx'
        wb.save(response)
        return response

class AppointmentViewSet(BaseModelViewSet):
    queryset = Appointment.objects.all()
    serializer_class = AppointmentSerializer
    http_method_names = ['get', 'post', 'put', 'delete']

    class StatsSerializer(serializers.Serializer):
        total_appointments = serializers.IntegerField()
        appointments_today = serializers.IntegerField()
        appointments_this_month = serializers.IntegerField()
        most_popular_service = serializers.CharField(allow_null=True)
        busiest_day = serializers.DateField(allow_null=True)

    @action(detail=False, methods=["GET"], url_path="stats")
    def get_stats(self, request):
        from django.utils import timezone
        from django.db.models.functions import TruncDate

        queryset = self.get_queryset()
        if not queryset.exists():
            return Response({
                'total_appointments': 0,
                'appointments_today': 0,
                'appointments_this_month': 0,
                'most_popular_service': None,
                'busiest_day': None
            })

        today = timezone.now().date()
        
        service_stats = queryset.values('service__name')\
            .annotate(count=Count('service'))\
            .order_by('-count').first()

        busiest_day = queryset.annotate(
            date_only=TruncDate('date')
        ).values('date_only')\
        .annotate(count=Count('id'))\
        .order_by('-count').first()

        stats = {
            'total_appointments': queryset.count(),
            'appointments_today': queryset.filter(date__date=today).count(),
            'appointments_this_month': queryset.filter(
                date__year=today.year,
                date__month=today.month
            ).count(),
            'most_popular_service': service_stats['service__name'] if service_stats else None,
            'busiest_day': busiest_day['date_only'] if busiest_day else None
        }
        
        serializer = self.StatsSerializer(instance=stats)
        return Response(serializer.data)

    def get_queryset(self):
        qs = super().get_queryset()
        if not self.request.user.is_superuser:
            return qs.filter(client__user=self.request.user)
        return qs

    def perform_create(self, serializer):
        # Берем первого клиента пользователя
        client = Client.objects.filter(user=self.request.user).first()
        if not client:
            raise serializers.ValidationError("Client profile not found for current user")
        serializer.save(client=client)

    @action(detail=False, methods=['GET'], url_path='export-excel')
    def export_excel(self, request):
        # Создаем новый excel файл
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Записи"

        # аголовки
        headers = ['Клиент', 'Услуга', 'Мастер', 'Дата', 'Время']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)

        # Данные
        queryset = self.get_queryset()
        for row, appointment in enumerate(queryset, 2):
            ws.cell(row=row, column=1, value=appointment.client.name)
            ws.cell(row=row, column=2, value=appointment.service.name)
            ws.cell(row=row, column=3, value=appointment.master.name)
            ws.cell(row=row, column=4, value=appointment.date.strftime('%d.%m.%Y'))
            ws.cell(row=row, column=5, value=appointment.date.strftime('%H:%M'))

        # Автоматическая ширина колонок
        for column in ws.columns:
            max_length = 0
            column_letter = openpyxl.utils.get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Создаем HTTP response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=appointments.xlsx'

        # Сохраняем файл
        wb.save(response)
        return response

class ReviewViewSet(BaseModelViewSet):
    queryset = Review.objects.all()
    serializer_class = ReviewSerializer
    http_method_names = ['get', 'post', 'put', 'delete']

    class StatsSerializer(serializers.Serializer):
        total_reviews = serializers.IntegerField()
        avg_rating = serializers.FloatField(allow_null=True)
        five_star_reviews = serializers.IntegerField()
        one_star_reviews = serializers.IntegerField()
        most_reviewed_service = serializers.CharField(allow_null=True)

    @action(detail=False, methods=["GET"], url_path="stats")
    def get_stats(self, request):
        queryset = self.get_queryset()
        if not queryset.exists():
            return Response({
                'total_reviews': 0,
                'avg_rating': None,
                'five_star_reviews': 0,
                'one_star_reviews': 0,
                'most_reviewed_service': None
            })

        service_stats = queryset.values('service__name')\
            .annotate(count=Count('service'))\
            .order_by('-count').first()

        stats = {
            'total_reviews': queryset.count(),
            'avg_rating': queryset.aggregate(Avg('rating'))['rating__avg'],
            'five_star_reviews': queryset.filter(rating=5).count(),
            'one_star_reviews': queryset.filter(rating=1).count(),
            'most_reviewed_service': service_stats['service__name'] if service_stats else None
        }
        
        serializer = self.StatsSerializer(instance=stats)
        return Response(serializer.data)

    def get_queryset(self):
        qs = super().get_queryset()
        if not self.request.user.is_superuser:
            return qs.filter(client__user=self.request.user)
        return qs

    def perform_create(self, serializer):
        try:
            # Используем filter().first() вместо get()
            client = Client.objects.filter(user=self.request.user).first()
            if not client:
                raise serializers.ValidationError("Client profile not found for current user")
            serializer.save(client=client)
        except Exception as e:
            raise serializers.ValidationError(str(e))

class UserViewSet(viewsets.ModelViewSet):
    permission_classes = [IsAuthenticated]
    
    @action(detail=False, methods=['GET'])
    def profile(self, request):
        user = request.user
        return Response({
            'id': user.id,
            'username': user.username,
            'email': user.email,
            'date_joined': user.date_joined
        })
